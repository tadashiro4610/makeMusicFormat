from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side
import openpyxl
from openpyxl.styles.alignment import Alignment
import tkinter as tk
import os

class msf:
    def __init__(self,musicName,barLow,section_entry_list,bar_entry_list,col16_list,dir):
        self.musicName=musicName
        self.barLow=int(barLow)
        self.section_entry_list=section_entry_list
        self.bar_entry_list=bar_entry_list
        self.col16_list=col16_list
        self.dir=dir
        
        self.wb=openpyxl.Workbook()
        self.wb.save(self.dir)
        self.ws=self.wb["Sheet"]

    def wright(self):
        ###書いたり塗ったり
        #i=行
        #x=書き込む行
        x=1
        for i in range(len(self.section_entry_list)):
            tmp=1
            #構成名
            compName=self.section_entry_list[i]
            #小節数
            bar=int(self.bar_entry_list[i])
            #カラーコード
            colNum=self.col16_list[i]

            #小節名
            self.ws.cell(row=x,column=1).value=compName
            #色のフォーマット
            fill = openpyxl.styles.PatternFill(patternType='solid',fgColor=colNum, bgColor=colNum)
            #罫線のフォーマット
            border = Border(top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000'), 
                        left=Side(style='thin', color='000000'),right=Side(style='thin', color='000000'))
            #小節数
            for j in range(2,bar+2):
                if tmp==1:
                    k=j
                    tmp=0
                if (j-2)%self.barLow==0 and j>self.barLow-1:
                    x=x+2
                    k=2
                self.ws.cell(row=x,column=k).value=j-1
                self.ws.cell(row=x,column=k).fill=fill
                self.ws.cell(row=x,column=k).border=border
                k=k+1
            x=x+3
        #題名記述(行を挿入して、セルを結合して、題名を書き込んで、中央揃え)
        self.ws.insert_rows(1,2)
        self.ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=self.barLow+2)
        self.ws.cell(1,1).value=self.musicName
        self.ws.cell(1,1).alignment=Alignment(horizontal='center',vertical='center')
        self.wb.save(self.dir)